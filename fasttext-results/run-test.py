import os
import glob
import openpyxl
import time
from fasttext_tools import fasttext_runner as ft_runner
from fasttext_tools import data_set_fixer_script as dsf
from openpyxl.styles import Alignment

data_set_path = '../WSD_Dataset/'
words = []

# Iterative over each file in WSD_Dataset and generate test and training data for each

paths_iter = iter(os.walk(data_set_path))
next(paths_iter) # skipping the first element in paths as that just lists all WSD_Dataset/ sub dirs  (the word dirs)
for path_data in paths_iter:
    full_path = path_data[0] + '/'
    word_name = full_path.split('/')[-2]
    words.append(word_name)
    # print('\t {} \t {}'.format(full_path, word_name))

    dsf.main(
        label_file_path=full_path+'classes_map.txt',
        key_file_path=full_path+'train.gold.txt',
        data_file_path=full_path+'train.data.txt',
        output_file='datasets/'+word_name+'-training-data.txt'
    )

    dsf.main(
        label_file_path=full_path+'classes_map.txt',
        key_file_path=full_path+'test.gold.txt',
        data_file_path=full_path+'test.data.txt',
        output_file='datasets/'+word_name+'-testing-data.txt'
    )

###########
# Convert training sets to one large file (train one model with all)

full_training_data = ""
for training_file in glob.glob('datasets/*-training-data.txt'):
    with open(training_file) as f:
        f_data  = f.read()
        full_training_data += f_data

with open("datasets-full/full-training-data.txt", 'w') as f:
    f.write(full_training_data)

###########

html_overall_table_body = ""
html_word_table_body = ""

wb = openpyxl.Workbook()
ws_results = wb.active
ws_results.title = "Results"
file_output_name = 'results.xlsx'
# model = ft_runner.get_model('datasets-full/full-training-data.txt') #######
row_count = 2
ws_results['A1'] = 'Word'
ws_results['B1'] = 'Accuracy'
ws_results['C1'] = 'F1-score (macro)'
ws_results['D1'] = 'Confidence'
for word in words:
    start_time_model_training = time.time()
    model = ft_runner.get_model('datasets/'+word+'-training-data.txt') ##########
    training_time = time.time() - start_time_model_training
    r = ft_runner.test_model(model,'datasets/'+word+'-testing-data.txt','results/'+word+'-results.json', training_data_path='datasets/'+word+'-training-data.txt')

    ws_results['A' + str(row_count)] = word
    ws_results['B' + str(row_count)] = r['percentage_correct']
    ws_results['C' + str(row_count)] = r['f1score (macro)']
    ws_results['D' + str(row_count)] = r['avg confidence']
    row_count += 1

    ws_word = wb.create_sheet(title=word)
    ws_word['A1'] = word
    ws_word.merge_cells('A1:B1')
    ws_word['A2'] = 'num_correct'
    ws_word['B2'] = r['num_correct']
    ws_word['A3'] = 'num_incorrect'
    ws_word['B3'] = r['num_incorrect']
    ws_word['A4'] = 'percentage_correct'
    ws_word['B4'] = r['percentage_correct']
    ws_word['A5'] = 'f1-score (macro)'
    ws_word['B5'] = r['f1score (macro)']
    ws_word['A6'] = 'f1-score (weighted)'
    ws_word['B6'] = r['f1score (weighted)']
    ws_word['A7'] = 'avg confidence'
    ws_word['B7'] = r['avg confidence']
    ws_word['A8'] = 'avg correct confidence'
    ws_word['B8'] = r['avg correct confidence']
    ws_word['A9'] = 'avg incorrect confidence'
    ws_word['B9'] = r['avg incorrect confidence']
    ws_word['A10'] = 'total'
    ws_word['B10'] = r['total']
    ws_word['A11'] = 'Avg num words (test data)'
    ws_word['B11'] = r['avg num words (test data)']
    ws_word['A12'] = 'Avg num words (training data)'
    ws_word['B12'] = r['avg num words (training data)']
    ws_word['A13'] = 'training time'
    ws_word['B13'] = training_time

    ws_word['E2'] = 'Label Data'
    ws_word.merge_cells('E2:F2')
    ws_word['E2'].alignment = Alignment(horizontal='center')
    label_row_count = 3
    for label in r['label_data']:
        ws_word['E' + str(label_row_count)] = label
        ws_word.merge_cells('E' + str(label_row_count)+':F' + str(label_row_count))
        ws_word['E' + str(label_row_count)].alignment = Alignment(horizontal='center')
        label_row_count += 1
        ws_word['E' + str(label_row_count)] = 'num_correct'
        ws_word['F' + str(label_row_count)] = r['label_data'][label]['num_correct']
        label_row_count +=1
        ws_word['E' + str(label_row_count)] = 'num_incorrect'
        ws_word['F' + str(label_row_count)] = r['label_data'][label]['num_incorrect']
        label_row_count += 1
        ws_word['E' + str(label_row_count)] = 'total'
        ws_word['F' + str(label_row_count)] = r['label_data'][label]['total']
        label_row_count += 1
        ws_word['E' + str(label_row_count)] = 'confidence'
        ws_word['F' + str(label_row_count)] = r['label_data'][label]['confidence']
        label_row_count += 1


#     print('{} : {}'.format(word, r['percentage_correct']))
#     html_overall_table_body += "<tr><td>{}</td><td>{}</td></tr>".format(word, r['percentage_correct'])
#     html_word_table_body += "<tr class='heading'><td colspan='3'>{}</td></tr><tr><td>num correct</td><td colspan='2'>{}</td></tr><tr><td>num incorrect</td><td colspan='2'>{}</td></tr><tr><td>total</td><td colspan='2'>{}</td></tr>".format(word,r['num_correct'], r['num_incorrect'], r['total'])
#     for label in r['label_data']:
#         if r['label_data'][label]['num_correct'] == 0:
#             html_word_table_body += "<tr class='bad'><td rowspan='3'>{}</td><td>num correct</td><td>{}</td></tr><tr><td>num incorrect</td><td>{}</td></tr><tr><td>total</td><td>{}</td></tr>".format(label, r['label_data'][label]['num_correct'], r['label_data'][label]['num_incorrect'], r['label_data'][label]['total'])
#         else:
#             html_word_table_body += "<tr><td rowspan='3'>{}</td><td>num correct</td><td>{}</td></tr><tr><td>num incorrect</td><td>{}</td></tr><tr><td>total</td><td>{}</td></tr>".format(
#                 label, r['label_data'][label]['num_correct'], r['label_data'][label]['num_incorrect'],
#                 r['label_data'][label]['total'])
#
# # build html
# html_head = "<html><head><title>Fasttext Results</title><style>.bad {color:red;} table, th, tr, td {border: 1px solid black;border-collapse: collapse;}.heading {font-weight: bold;align: centre;}</style></head>"
# html_body_pt1 = "<body><h3>Overall</h3><table><tr class='heading'><td>Word</td><td>Accuracy</td></tr>{}</table></body>".format(html_overall_table_body)
# html_body_pt2 = "<h3>Labels</h3><table>{}</table></html>".format(html_word_table_body)
#
# with open('results.html', 'w') as f:
#     f.write(html_head+html_body_pt1+html_body_pt2)
wb.save(filename=file_output_name)